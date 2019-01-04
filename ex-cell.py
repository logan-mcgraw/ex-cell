 #Logan McGraw
#Ex-cell, 1/3/19

import openpyxl, time
from openpyxl import load_workbook
from openpyxl import Workbook

#open the desired spreadsheets
filename1 = input('Spreadsheet to write from: \n') #this needs to include the file extension (.xlsx, .xlsm, etc)
wb1 = load_workbook(filename1)
source = wb1.active

filename2 = input('Spreadsheet to write to: \n')
wb2 = load_workbook(filename2)
sheet = wb2.active

#set up empty data lists(data specific to my sheets)
sampleName = []
pressure = []
thickness = []
thermCond = []
RA = []

#columns containing the desired data in the source spreadsheet
name = source['A']
p = source['F']
thicc = source['E']
TC = source['H']
ra = source['C']

#new lists help for iteration later
columns = [name, p, thicc, TC, ra]
lists = [sampleName, pressure, thickness, thermCond, RA]
cs = ['A', 'C', 'D', 'E', 'F']

n = 0

#move data from source sheet to lists
while n <= len(cs)-1: 
    for i in columns[n]:
        lists[n] += [i.value]
    del lists[n][0] #removes title from top of data
    n += 1
    

#finds first empty row, prevents overwriting existing data in master sheet
rowNum = 1
test = sheet['A' + str(rowNum)] #format 'A1' for example
while test.value != None: #goes down the column until an empty cell (with value None) is found
    rowNum += 1
    test = sheet['A' + str(rowNum)] 
    if test.value == None: 
        x = rowNum
        break

#write the data to the new sheet
n = 0
while n <= 4:
    for i in lists[n]:
        cell = cs[n] + str(x)
        dateCell = 'B' + str(x) #had to add date manually
        sheet[cell] = i
        sheet[dateCell] = time.strftime('%x', time.gmtime()) #prints date in format mm/dd/yy
        x += 1       
    n += 1
    x = rowNum

print('Data copied from', filename1, 'to', filename2)
        
wb2.save(filename2) #saves the master sheet
