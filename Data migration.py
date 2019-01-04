import openpyxl, time
from openpyxl import load_workbook
from openpyxl import Workbook

filename1 = input('Spreadsheet to write from     ') + str('.xlsx')
wb1 = load_workbook(filename1)
source = wb1.active

filename2 = input('Spreadsheet to write to     ') + str('.xlsx') #replace with master sheet name
wb2 = load_workbook(filename2)
sheet = wb2.active

#data specific to my sheets, change if using other sheets
sampleName = []
pressure = []
thickness = []
thermCond = []
RA = []

name = source['A']
p = source['F']
thicc = source['E']
TC = source['H']
ra = source['C']

columns = [name, p, thicc, TC, ra]
lists = [sampleName, pressure, thickness, thermCond, RA]
cs = ['A', 'C', 'D', 'E', 'F']

n = 0

#move data from source sheet to lists
while n <= 4:
    for i in columns[n]:
        lists[n] += [i.value]
    del lists[n][0] #removes title from top of data
    n += 1
    
print()

#alg to find first empty row
p = 1
test = sheet['A' + str(p)]

#this loop goes down a column in the sheet until it finds a blank one to start writing data
#prevents overwriting of previous data
while test.value != None:
    p += 1
    test = sheet['A' + str(p)]
    if test.value == None:
        x = p
        break
print('Starting on row', x)

#write the data to the new sheet
n = 0
while n <= 4:
    for i in lists[n]:
        cell = cs[n] + str(x)
        dateCell = 'B' + str(x)
        sheet[cell] = i
        sheet[dateCell] = time.strftime('%x', time.gmtime())
        x += 1       
    n += 1
    x = p

time.sleep(3) #gives the user the sense that something important is happening
print('Data copied from', filename1, 'to', filename2)
input()

        
wb2.save(filename2)

        
    
